import math
import os
from typing import List, Dict, Optional, Tuple, Set
from datetime import datetime
from openpyxl import Workbook
from enum import Enum

# Constants and Configuration
class TaskConstants:
    """Centralized constants for task management"""
    DEFAULT_VALUE = 3.0
    DEFAULT_TIME = 4.0
    MIN_VALUE = 0.0
    MAX_VALUE = 6.0
    MIN_TIME = 0.1
    MAX_TIME = 8.0
    MAX_DISPLAY_TASKS = 15
    
    # Priority score thresholds
    HIGH_PRIORITY_THRESHOLD = 2.0
    MEDIUM_PRIORITY_THRESHOLD = 1.0
    
    # Color schemes for task states
    COLOR_MOVED = '#2a82da'      # Blue  
    COLOR_ORIGINAL = '#e74c3c'   # Red
    COLOR_NEW = '#00FF7F'        # Spring Green - for newly added tasks

class TaskState(Enum):
    """Enumeration for task visual states"""
    ORIGINAL = "original"
    MOVED = "moved"
    NEW = "new"

class Task:
    def __init__(self, task: str, value: float, time: float, is_new: bool = False):
        self.task = task
        self.value = value
        self.time = time
        self.score = 0.0
        self.is_new = is_new  # Track if this is a newly added task

    def calculate_score(self):
        self.score = self.value / math.log(max(2.718, self.time))
        return self.score
    
    def mark_as_seen(self):
        """Mark task as no longer new"""
        self.is_new = False
    
    def get_state(self, moved_tasks_indices: Set[int], new_task_indices: Set[int], task_index: int) -> TaskState:
        """Determine the visual state of this task"""
        if task_index in new_task_indices or self.is_new:
            return TaskState.NEW
        elif task_index in moved_tasks_indices:
            return TaskState.MOVED
        else:
            return TaskState.ORIGINAL
    
    def get_color(self, moved_tasks_indices: Set[int], new_task_indices: Set[int], task_index: int) -> str:
        """Get the color for this task based on its state"""
        state = self.get_state(moved_tasks_indices, new_task_indices, task_index)
        color_map = {
            TaskState.MOVED: TaskConstants.COLOR_MOVED,
            TaskState.ORIGINAL: TaskConstants.COLOR_ORIGINAL,
            TaskState.NEW: TaskConstants.COLOR_NEW
        }
        return color_map[state]

class TaskStateManager:
    """Manages task states, highlighting, and visual tracking"""
    
    def __init__(self):
        self.moved_points: Set[int] = set()
        self.new_task_indices: Set[int] = set()
        self.highlighted_task_index: Optional[int] = None
    
    def mark_task_moved(self, task_index: int):
        """Mark a task as having been moved"""
        self.moved_points.add(task_index)
        # When a task is moved, it's no longer "new"
        self.new_task_indices.discard(task_index)
    
    def is_task_moved(self, task_index: int) -> bool:
        """Check if a task has been moved"""
        return task_index in self.moved_points
    
    def mark_task_new(self, task_index: int):
        """Mark a task as newly added"""
        self.new_task_indices.add(task_index)
    
    def is_task_new(self, task_index: int) -> bool:
        """Check if a task is new"""
        return task_index in self.new_task_indices
    
    def clear_new_tasks(self):
        """Clear all new task tracking"""
        self.new_task_indices.clear()
    
    def clear_moved_tasks(self):
        """Clear all moved task tracking"""
        self.moved_points.clear()
    
    def set_highlighted_task(self, task_index: Optional[int]):
        """Set which task is currently highlighted"""
        self.highlighted_task_index = task_index
    
    def get_highlighted_task(self) -> Optional[int]:
        """Get the currently highlighted task index"""
        return self.highlighted_task_index
    
    def clear_highlighting(self):
        """Clear task highlighting"""
        self.highlighted_task_index = None

class TaskDisplayFormatter:
    """Handles formatting tasks for display in UI components"""
    
    @staticmethod
    def format_rank(rank: int) -> str:
        """Format task rank with medals for top 3"""
        if rank == 1:
            return "🥇"  # Gold medal for #1
        elif rank == 2:
            return "🥈"  # Silver medal for #2
        elif rank == 3:
            return "🥉"  # Bronze medal for #3
        else:
            return f"#{rank}"
    
    @staticmethod
    def format_priority_score(score: float) -> str:
        """Format priority score with visual indicators"""
        if score >= TaskConstants.HIGH_PRIORITY_THRESHOLD:
            return f"🔥{score:.2f}"  # High priority
        elif score >= TaskConstants.MEDIUM_PRIORITY_THRESHOLD:
            return f"⚡{score:.2f}"  # Medium priority
        else:
            return f"💡{score:.2f}"  # Lower priority
    
    @staticmethod
    def format_task_name(task: Task) -> str:
        """Format task name"""
        return task.task
    
    @staticmethod
    def format_value(value: float) -> str:
        """Format value with star indicator"""
        return f"{value:.1f}⭐"
    
    @staticmethod
    def get_tooltip_text(task: Task) -> str:
        """Generate tooltip text for a task"""
        return f"Full task: {task.task}"

class TaskValidator:
    """Validates task data and provides default values"""
    
    @staticmethod
    def validate_value(value: float) -> bool:
        """Validate task value is within acceptable range"""
        return TaskConstants.MIN_VALUE <= value <= TaskConstants.MAX_VALUE
    
    @staticmethod
    def validate_time(time: float) -> bool:
        """Validate task time is within acceptable range"""
        return TaskConstants.MIN_TIME <= time <= TaskConstants.MAX_TIME
    
    @staticmethod
    def validate_task_name(name: str) -> bool:
        """Validate task name is not empty"""
        return bool(name and name.strip())
    
    @staticmethod
    def get_default_values() -> Tuple[float, float]:
        """Get default value and time for new tasks"""
        return TaskConstants.DEFAULT_VALUE, TaskConstants.DEFAULT_TIME
    
    @staticmethod
    def sanitize_task_name(name: str) -> str:
        """Clean and sanitize task name"""
        return name.strip()
    
    @staticmethod
    def create_validated_task(name: str, value: float = None, time: float = None) -> Task:
        """Create a task with validated inputs"""
        # Sanitize name
        clean_name = TaskValidator.sanitize_task_name(name)
        if not TaskValidator.validate_task_name(clean_name):
            raise ValueError("Task name cannot be empty")
        
        # Use defaults if not provided
        if value is None:
            value = TaskConstants.DEFAULT_VALUE
        if time is None:
            time = TaskConstants.DEFAULT_TIME
        
        # Validate ranges
        if not TaskValidator.validate_value(value):
            raise ValueError(f"Value must be between {TaskConstants.MIN_VALUE} and {TaskConstants.MAX_VALUE}")
        if not TaskValidator.validate_time(time):
            raise ValueError(f"Time must be between {TaskConstants.MIN_TIME} and {TaskConstants.MAX_TIME}")
        
        return Task(clean_name, value, time)

class SampleDataGenerator:
    """Generates sample task data for testing and demonstration"""
    
    @staticmethod
    def get_sample_tasks() -> List[Task]:
        """Generate realistic sample tasks for testing"""
        sample_data = [
            ("Complete Project Proposal", 4.5, 3.0),
            ("Review Code Changes", 3.0, 2.0),
            ("Team Meeting", 2.5, 1.5),
            ("Update Documentation", 3.5, 4.0),
            ("Bug Fixing", 4.0, 2.5),
            ("Client Presentation", 5.0, 4.0),
            ("Code Refactoring", 3.5, 5.0),
            ("Unit Testing", 4.0, 3.0),
            ("Performance Optimization", 4.5, 6.0),
            ("Security Audit", 5.0, 4.5),
            ("Database Migration", 4.0, 7.0),
            ("API Integration", 3.5, 3.5),
            ("User Training", 3.0, 2.0),
            ("System Backup", 2.5, 1.0),
            ("Deployment Planning", 4.0, 2.0),
            ("Code Review", 3.5, 1.5),
            ("Feature Implementation", 4.5, 5.0),
            ("Technical Documentation", 3.0, 4.0),
            ("Bug Triage", 3.5, 2.0),
            ("System Monitoring", 2.5, 1.5)
        ]
        
        return [Task(name, value, time) for name, value, time in sample_data]
    
    @staticmethod
    def create_tasks_from_text(text: str) -> List[Task]:
        """Create tasks from clipboard or text input"""
        if not text or not text.strip():
            return []
        
        lines = [line.strip() for line in text.split('\n') if line.strip()]
        tasks = []
        
        for line in lines:
            try:
                task = TaskValidator.create_validated_task(line)
                tasks.append(task)
            except ValueError:
                # Skip invalid lines
                continue
        
        return tasks
    
    @staticmethod
    def create_tasks_from_mindmap(text: str) -> List[Task]:
        """Create tasks from mindmap-style indented text
        
        Example input:
        dette er en tekst
            dette er een annen tekst
                dette er en fjerde tekst
                dette er en femte tekst
            dette er en tredje tekst
        
        Creates tasks like:
        - dette er en tekst->dette er een annen tekst
        - dette er een annen tekst->dette er en fjerde tekst
        - etc.
        """
        if not text or not text.strip():
            return []
        
        lines = text.split('\n')
        tasks = []
        parent_stack = []  # Stack to keep track of parent nodes at each level
        
        for line in lines:
            if not line.strip():
                continue
                
            # Count indentation level (4 spaces = 1 level)
            stripped_line = line.lstrip()
            indent_level = (len(line) - len(stripped_line)) // 4
            content = stripped_line.strip()
            
            if not content:
                continue
            
            # Adjust parent stack to current level
            while len(parent_stack) > indent_level:
                parent_stack.pop()
            
            # If we have a parent, create a task relationship
            if parent_stack:
                parent_content = parent_stack[-1]
                task_name = f"{parent_content}->{content}"
                try:
                    task = TaskValidator.create_validated_task(task_name)
                    tasks.append(task)
                except ValueError:
                    # Skip invalid task names
                    continue
            
            # Add current content to parent stack
            if len(parent_stack) == indent_level:
                parent_stack.append(content)
            else:
                # Fill gaps in hierarchy if needed
                while len(parent_stack) < indent_level:
                    parent_stack.append("")
                parent_stack.append(content)
        
        return tasks

class ExcelExporter:
    """Handles Excel export functionality"""
    
    @staticmethod
    def export_tasks_to_excel(tasks: List[Task], file_path: str) -> bool:
        """Export tasks to Excel file"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Task Priorities"
            
            # Add headers
            headers = ['📋 Task', '★ Value', '⏰ Time (hours)', '🏆 Priority Score']
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Add data
            sorted_tasks = calculate_and_sort_tasks(tasks)
            for row, task in enumerate(sorted_tasks, 2):
                ws.cell(row=row, column=1, value=task.task)
                ws.cell(row=row, column=2, value=task.value)
                ws.cell(row=row, column=3, value=task.time)
                ws.cell(row=row, column=4, value=task.score)
            
            # Format columns
            for col in range(1, 5):
                ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 15
            
            # Save the file
            wb.save(file_path)
            return True
            
        except Exception as e:
            print(f"Export error: {e}")
            return False
    
    @staticmethod
    def get_default_export_path() -> str:
        """Get default export path"""
        save_locations = [
            os.path.expanduser("~/Downloads"),
            os.path.expanduser("~/Desktop"),
            os.path.expanduser("~"),
            os.getcwd()
        ]
        
        for location in save_locations:
            if os.path.exists(location) and os.access(location, os.W_OK):
                return location
        
        return os.getcwd()  # Last resort
    
    @staticmethod
    def generate_filename() -> str:
        """Generate a timestamped filename"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        return f"priority_analysis_{timestamp}.xlsx"

def calculate_and_sort_tasks(tasks: List[Task]) -> List[Task]:
    for t in tasks:
        t.calculate_score()
    return sorted(tasks, key=lambda t: t.score, reverse=True)

def get_top_tasks(tasks: List[Task], count: int = 3) -> List[Task]:
    """Get the top N tasks by priority score"""
    sorted_tasks = calculate_and_sort_tasks(tasks)
    return sorted_tasks[:count]

def get_task_colors(tasks: List[Task], moved_indices: Set[int], new_task_indices: Set[int] = None) -> List[str]:
    """Get colors for all tasks based on their states"""
    if new_task_indices is None:
        new_task_indices = set()
    colors = []
    for i, task in enumerate(tasks):
        colors.append(task.get_color(moved_indices, new_task_indices, i))
    return colors 