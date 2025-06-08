from PyQt6.QtWidgets import (QWidget, QVBoxLayout, QPushButton, QTabWidget, QTableWidget, 
                             QTableWidgetItem, QLineEdit, QLabel, QHBoxLayout, QMessageBox, 
                             QFileDialog, QSplitter, QFrame, QCalendarWidget, QTimeEdit, 
                             QListWidget, QListWidgetItem, QGroupBox, QFormLayout, QAbstractItemView,
                             QDialog, QDialogButtonBox)
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
from PyQt6.QtCore import Qt, QTimer, QDate, QTime, QMimeData, QPoint
from PyQt6.QtGui import QClipboard, QColor, QDrag
from .model import Task, calculate_and_sort_tasks
import numpy as np
from openpyxl import Workbook
from datetime import datetime
from PyQt6.QtWidgets import QApplication

class ScheduleCalendarWidget(QCalendarWidget):
    """Custom calendar widget that shows days with scheduled tasks in bold"""
    
    def __init__(self, parent_widget, parent=None):
        super().__init__(parent)
        self.parent_widget = parent_widget
        self.highlighted_date = None  # Track currently highlighted date during drag
        self.drag_in_progress = False  # Track if we're currently in a drag operation
        
        # Timer to clean up highlighting if drag events don't fire properly
        self.highlight_cleanup_timer = QTimer()
        self.highlight_cleanup_timer.setSingleShot(True)
        self.highlight_cleanup_timer.timeout.connect(self.clear_drop_highlighting)
        
        # Prevent month changes during drag operations
        self.setVerticalHeaderFormat(QCalendarWidget.VerticalHeaderFormat.NoVerticalHeader)
        self.setNavigationBarVisible(True)  # Keep navigation visible but disable during drag
    
    def paintCell(self, painter, rect, date):
        """Override to paint cells with scheduled tasks in bold"""
        # Check if this date has any scheduled tasks
        has_tasks = self.has_scheduled_tasks_on_date(date)
        
        if has_tasks:
            # Get the current text format
            format = self.dateTextFormat(date)
            
            # Make it bold and change color slightly
            font = format.font()
            font.setBold(True)
            format.setFont(font)
            
            # Optional: add a subtle background color
            format.setBackground(QColor(42, 130, 218, 40))  # Light blue background
            
            # Apply the format
            self.setDateTextFormat(date, format)
        else:
            # Reset to normal format
            format = self.dateTextFormat(date)
            font = format.font()
            font.setBold(False)
            format.setFont(font)
            format.setBackground(QColor())  # Clear background
            self.setDateTextFormat(date, format)
        
        # Call parent to do the actual painting
        super().paintCell(painter, rect, date)
    
    def has_scheduled_tasks_on_date(self, qdate):
        """Check if any tasks are scheduled on the given QDate"""
        if not hasattr(self.parent_widget, 'task_list') or not self.parent_widget.task_list:
            return False
        
        # Convert QDate to Python date for comparison
        target_date = datetime(qdate.year(), qdate.month(), qdate.day()).date()
        
        for task in self.parent_widget.task_list:
            if task.is_scheduled() and task.scheduled_date.date() == target_date:
                return True
        
        return False
    
    def refresh_calendar_display(self):
        """Force refresh of the calendar to update bold formatting"""
        # Trigger a repaint by changing the selection slightly
        current = self.selectedDate()
        self.setSelectedDate(current)
        self.update()

    def highlight_date_for_drop(self, target_date):
        """Highlight a specific date during drag operations"""
        # Clear previous highlighting
        self.clear_drop_highlighting()
        
        # Convert Python date to QDate
        if isinstance(target_date, datetime):
            target_date = target_date.date()
        
        q_date = QDate(target_date.year, target_date.month, target_date.day)
        
        # Store the highlighted date
        self.highlighted_date = q_date
        
        # Create highlighting format
        highlight_format = self.dateTextFormat(q_date)
        highlight_format.setBackground(QColor(255, 20, 20, 255))  # Bright red with full opacity - impossible to miss!
        highlight_format.setForeground(QColor(255, 255, 255, 255))  # Pure white text with full opacity
        
        # Make it bold to stand out more
        font = highlight_format.font()
        font.setBold(True)
        font.setWeight(900)  # Maximum boldness
        font.setPointSize(font.pointSize() + 3)  # Make text much larger
        highlight_format.setFont(font)
        
        # Remove underline to avoid enum issues - the bright red background should be enough!
        # highlight_format.setUnderlineStyle(1)  # Single underline
        # highlight_format.setUnderlineColor(QColor(255, 255, 255))  # White underline
        
        # Apply the highlighting
        self.setDateTextFormat(q_date, highlight_format)
        
        # ALSO temporarily change the calendar selection for additional visual feedback
        current_selection = self.selectedDate()
        if current_selection != q_date:
            # Store the original selection to restore later
            if not hasattr(self, 'original_selection'):
                self.original_selection = current_selection
            # Change selection to the highlighted date for additional visual feedback
            self.setSelectedDate(q_date)
        
        # Debug message to confirm the method is being called
        print(f"🎯 Highlighting date: {q_date.toString()} with bright red background")
        
        # Force a refresh to show the highlighting immediately
        self.update()
        
        # Start cleanup timer as fallback (clear after 5 seconds of inactivity)
        self.highlight_cleanup_timer.start(5000)
    
    def clear_drop_highlighting(self):
        """Clear any drag drop highlighting"""
        # Stop the cleanup timer
        if hasattr(self, 'highlight_cleanup_timer'):
            self.highlight_cleanup_timer.stop()
            
        if self.highlighted_date:
            print(f"🧹 Clearing highlighting for date: {self.highlighted_date.toString()}")
            
            # First, completely clear the date format to remove any custom styling
            self.setDateTextFormat(self.highlighted_date, self.dateTextFormat(QDate()))
            
            # Then check if this date should have special formatting (scheduled tasks)
            target_date = datetime(self.highlighted_date.year(), self.highlighted_date.month(), self.highlighted_date.day()).date()
            if self.has_scheduled_tasks_on_date(self.highlighted_date):
                # Apply scheduled task formatting
                scheduled_format = self.dateTextFormat(QDate())  # Start with clean format
                font = scheduled_format.font()
                font.setBold(True)
                scheduled_format.setFont(font)
                scheduled_format.setBackground(QColor(42, 130, 218, 40))  # Blue background for scheduled
                self.setDateTextFormat(self.highlighted_date, scheduled_format)
            
            # Restore original calendar selection if we changed it
            if hasattr(self, 'original_selection'):
                self.setSelectedDate(self.original_selection)
                delattr(self, 'original_selection')
            
            self.highlighted_date = None
            
            # Force a complete refresh of the calendar
            self.updateCells()
            self.update()

class TimeSelectionDialog(QDialog):
    """Dialog for selecting start and end times when scheduling a task"""
    
    def __init__(self, task_name, selected_date, parent=None):
        super().__init__(parent)
        self.setWindowTitle(f"⏰ Schedule: {task_name}")
        self.setModal(True)
        self.resize(300, 200)
        
        # Apply dark theme
        self.setStyleSheet("""
            QDialog {
                background-color: #353535;
                color: white;
            }
            QLabel {
                color: white;
                font-size: 13px;
            }
            QTimeEdit {
                background-color: #555555;
                color: white;
                border: 1px solid #666666;
                border-radius: 4px;
                padding: 6px;
                font-size: 12px;
            }
            QPushButton {
                background-color: #2a82da;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 4px;
                font-weight: 500;
                min-width: 80px;
            }
            QPushButton:hover {
                background-color: #3292ea;
            }
            QPushButton:pressed {
                background-color: #1a72ca;
            }
        """)
        
        layout = QVBoxLayout()
        
        # Date display
        date_label = QLabel(f"📅 Date: {selected_date.strftime('%A, %B %d, %Y')}")
        date_label.setStyleSheet("font-weight: bold; margin-bottom: 10px; color: #2a82da;")
        layout.addWidget(date_label)
        
        # Task name display
        task_label = QLabel(f"📋 Task: {task_name}")
        task_label.setStyleSheet("margin-bottom: 15px;")
        layout.addWidget(task_label)
        
        # Time selection form
        form_layout = QFormLayout()
        
        self.start_time = QTimeEdit()
        self.start_time.setTime(QTime(9, 0))  # Default 9:00 AM
        self.start_time.setDisplayFormat("hh:mm AP")
        
        self.end_time = QTimeEdit()
        self.end_time.setTime(QTime(10, 0))  # Default 10:00 AM  
        self.end_time.setDisplayFormat("hh:mm AP")
        
        form_layout.addRow("🕐 Start Time:", self.start_time)
        form_layout.addRow("🕐 End Time:", self.end_time)
        
        layout.addLayout(form_layout)
        
        # Buttons
        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)
        
        self.setLayout(layout)
    
    def get_times(self):
        """Return the selected start and end times as strings"""
        return (self.start_time.time().toString("hh:mm AP"),
                self.end_time.time().toString("hh:mm AP"))

class DraggableTableWidget(QTableWidget):
    """Custom table widget that supports dragging tasks to calendar"""
    
    def __init__(self, parent_widget, parent=None):
        super().__init__(parent)
        self.parent_widget = parent_widget
    
    def startDrag(self, supportedActions):
        """Override to provide custom MIME data for task dragging"""
        item = self.currentItem()
        if item and item.row() >= 0:
            # Get the task index from the current row
            row = item.row()
            
            # Find the actual task index in the sorted list
            if hasattr(self.parent_widget, 'task_list') and self.parent_widget.task_list:
                # Calculate scores and sort tasks to match the table order
                for task in self.parent_widget.task_list:
                    task.calculate_score()
                sorted_tasks = sorted(self.parent_widget.task_list, key=lambda t: t.score, reverse=True)
                
                if row < len(sorted_tasks):
                    # Find the original index of this task
                    selected_task = sorted_tasks[row]
                    original_index = self.parent_widget.task_list.index(selected_task)
                    
                    # Create drag operation
                    drag = QDrag(self)
                    mimeData = QMimeData()
                    mimeData.setText(f"task_{original_index}")
                    drag.setMimeData(mimeData)
                    
                    # Execute drag
                    drag.exec(supportedActions)

class PriorityPlotWidget(QWidget):
    def __init__(self, task_list=None):
        super().__init__()
        self.task_list = task_list if task_list is not None else []
        self.dragging = False
        self.drag_index = None
        self.moved_points = set()  # Track which points have been moved
        self.current_annotation = None  # Track current hover annotation
        self.auto_update_timer = QTimer()  # Timer for real-time updates
        self.auto_update_timer.timeout.connect(self.update_priority_display)
        self.auto_update_timer.setSingleShot(True)
        self.initUI()
        
        # Show welcome message for new users
        if not self.task_list:
            self.show_welcome_message()

    def initUI(self):
        layout = QVBoxLayout()
        
        # Create main splitter for side-by-side layout when tasks exist
        self.main_splitter = QSplitter(Qt.Orientation.Horizontal)
        
        # Left panel for input (will be hidden after tasks are added)
        self.input_panel = QWidget()
        self.input_panel.setMaximumWidth(400)
        self.input_panel.setMinimumWidth(350)
        
        # Right panel for plot and results
        self.plot_panel = QWidget()
        
        self.main_splitter.addWidget(self.input_panel)
        self.main_splitter.addWidget(self.plot_panel)
        
        # Initially show only input panel
        self.plot_panel.hide()
        
        layout.addWidget(self.main_splitter)
        self.setLayout(layout)
        
        self.initInputPanel()
        self.initPlotPanel()

    def initInputPanel(self):
        layout = QVBoxLayout()
        
        # Streamlined header
        header_label = QLabel("🎯 Add Your Tasks")
        header_label.setStyleSheet("color: #ffffff; font-weight: bold; font-size: 16px; padding: 15px 10px 10px 10px;")
        layout.addWidget(header_label)
        
        # Quick start options in a more prominent layout
        quick_start_frame = QFrame()
        quick_start_frame.setStyleSheet("""
            QFrame {
                background-color: #404040;
                border-radius: 8px;
                padding: 10px;
                margin: 5px;
            }
        """)
        quick_layout = QVBoxLayout()
        
        quick_label = QLabel("🚀 Quick Start:")
        quick_label.setStyleSheet("color: #ffffff; font-weight: bold; margin-bottom: 8px;")
        quick_layout.addWidget(quick_label)
        
        # Horizontal layout for quick start buttons
        quick_buttons_layout = QHBoxLayout()
        
        # Enhanced test goals button (most prominent)
        self.test_button = QPushButton("🧪 Try Sample Tasks")
        self.test_button.clicked.connect(self.add_test_goals_and_proceed)
        self.test_button.setToolTip("🚀 Instantly try the app with 20 realistic work tasks!")
        self.test_button.setStyleSheet("""
            QPushButton {
                background-color: #28a745;
                font-weight: bold;
                font-size: 13px;
                padding: 12px;
                border-radius: 6px;
            }
            QPushButton:hover {
                background-color: #34ce57;
            }
        """)
        quick_buttons_layout.addWidget(self.test_button)
        
        # Enhanced clipboard import button
        self.clipboard_button = QPushButton("📋 Import List")
        self.clipboard_button.clicked.connect(self.add_goals_from_clipboard_and_proceed)
        self.clipboard_button.setToolTip("📄 Paste a list of tasks from your clipboard!")
        self.clipboard_button.setStyleSheet("""
            QPushButton {
                background-color: #007bff;
                font-weight: bold;
                font-size: 13px;
                padding: 12px;
                border-radius: 6px;
            }
            QPushButton:hover {
                background-color: #0056b3;
            }
        """)
        quick_buttons_layout.addWidget(self.clipboard_button)
        
        quick_layout.addLayout(quick_buttons_layout)
        quick_start_frame.setLayout(quick_layout)
        layout.addWidget(quick_start_frame)
        
        # Separator
        separator = QLabel("─── or add manually ───")
        separator.setStyleSheet("color: #888888; text-align: center; margin: 15px 0px;")
        separator.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(separator)
        
        # Simplified manual input
        form_layout = QVBoxLayout()
        
        self.task_input = QLineEdit()
        self.task_input.setPlaceholderText("Type a task and press Enter...")
        self.task_input.setToolTip("💡 Type a task name and press Enter to add it quickly!")
        self.task_input.returnPressed.connect(self.add_task_and_auto_proceed)
        self.task_input.setStyleSheet("""
            QLineEdit {
                padding: 12px;
                font-size: 14px;
                border-radius: 6px;
                border: 2px solid #555555;
            }
            QLineEdit:focus {
                border: 2px solid #2a82da;
            }
        """)
        form_layout.addWidget(self.task_input)
        
        # Add button (less prominent since Enter is preferred)
        self.add_button = QPushButton("➕ Add Task")
        self.add_button.clicked.connect(self.add_task_and_auto_proceed)
        self.add_button.setStyleSheet("""
            QPushButton {
                background-color: #6c757d;
                padding: 8px;
                border-radius: 4px;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #5a6268;
            }
        """)
        form_layout.addWidget(self.add_button)
        
        layout.addLayout(form_layout)
        
        # Task list (more compact)
        list_label = QLabel("📋 Your Tasks:")
        list_label.setStyleSheet("color: #ffffff; font-weight: bold; margin-top: 20px; margin-bottom: 5px;")
        layout.addWidget(list_label)
        
        self.input_table = QTableWidget()
        self.input_table.setColumnCount(1)
        self.input_table.setHorizontalHeaderLabels(["Task"])
        self.input_table.setMaximumHeight(200)  # Limit height
        self.input_table.setStyleSheet("""
            QTableWidget {
                border-radius: 4px;
                font-size: 12px;
            }
        """)
        layout.addWidget(self.input_table)
        
        # Help button (smaller, less prominent)
        help_button = QPushButton("❓ Help")
        help_button.clicked.connect(self.show_help)
        help_button.setStyleSheet("""
            QPushButton {
                background-color: #6c757d;
                font-size: 11px;
                padding: 6px 12px;
                border-radius: 4px;
                margin-top: 10px;
            }
            QPushButton:hover {
                background-color: #5a6268;
            }
        """)
        layout.addWidget(help_button)
        
        layout.addStretch()
        self.input_panel.setLayout(layout)
        self.refresh_input_table()

    def initPlotPanel(self):
        layout = QVBoxLayout()
        
        # Header with real-time priority info
        self.priority_header = QLabel("🎯 Drag tasks to prioritize • Drag from priority table to calendar to schedule • Top 3 priorities shown below")
        self.priority_header.setStyleSheet("color: #ffffff; font-weight: bold; padding: 10px; font-size: 14px;")
        layout.addWidget(self.priority_header)
        
        # Create horizontal splitter for chart and calendar side by side
        main_splitter = QSplitter(Qt.Orientation.Horizontal)
        
        # Left side: Plot and live results
        left_panel = QWidget()
        left_layout = QVBoxLayout()
        
        # Create splitter for plot and live results
        plot_splitter = QSplitter(Qt.Orientation.Vertical)
        
        # Plot area
        plot_widget = QWidget()
        plot_layout = QVBoxLayout()
        
        self.figure = Figure(figsize=(6, 4), facecolor='#353535')
        self.canvas = FigureCanvas(self.figure)
        self.ax = self.figure.add_subplot(111)
        
        # Set modern plot style
        self.ax.set_facecolor('#353535')
        self.ax.grid(True, linestyle='--', alpha=0.3, color='#555555')
        self.ax.spines['bottom'].set_color('#555555')
        self.ax.spines['top'].set_color('#555555')
        self.ax.spines['left'].set_color('#555555')
        self.ax.spines['right'].set_color('#555555')
        
        # Set labels with modern styling
        self.ax.set_xlabel('* Value (Impact/Importance)', color='white', fontsize=11, fontweight='bold')
        self.ax.set_ylabel('Time Investment (Hours)', color='white', fontsize=11, fontweight='bold')
        self.ax.set_title('Interactive Priority Matrix', color='white', fontsize=13, fontweight='bold', pad=15)
        
        # Style the ticks
        self.ax.tick_params(colors='white', which='both')
        
        # Set fixed axis limits
        self.ax.set_xlim(0, 6)
        self.ax.set_ylim(0, 8)
        
        # Connect mouse events
        self.canvas.mpl_connect('button_press_event', self.on_press)
        self.canvas.mpl_connect('button_release_event', self.on_release)
        self.canvas.mpl_connect('motion_notify_event', self.on_motion)
        self.canvas.mpl_connect('motion_notify_event', self.on_hover)
        
        plot_layout.addWidget(self.canvas)
        plot_widget.setLayout(plot_layout)
        
        # Live results area
        results_widget = QWidget()
        results_widget.setMinimumHeight(300)  # Increased from 200px for better visibility
        results_widget.setMaximumHeight(500)  # Allow more room to grow
        results_layout = QVBoxLayout()
        
        # Live priority list header - make it more prominent
        live_header = QLabel("🏆 Live Priority Ranking (Drag to Calendar →)")
        live_header.setStyleSheet("""
            color: #ffffff; 
            font-weight: bold; 
            font-size: 16px; 
            padding: 8px; 
            background-color: #2a82da; 
            border-radius: 6px; 
            margin-bottom: 5px;
        """)
        results_layout.addWidget(live_header)
        
        # Compact results table with much better readability
        self.live_table = DraggableTableWidget(self)
        self.live_table.setColumnCount(4)
        self.live_table.setHorizontalHeaderLabels(['🏆', 'Task', 'Value', 'Score'])
        
        # Dramatically improved styling for better readability
        self.live_table.setStyleSheet("""
            QTableWidget {
                font-size: 14px;
                border-radius: 6px;
                border: 2px solid #555555;
                background-color: #404040;
                selection-background-color: #2a82da;
                gridline-color: #666666;
            }
            QTableWidget::item {
                padding: 12px 8px;
                border-bottom: 1px solid #555555;
                min-height: 16px;
            }
            QTableWidget::item:selected {
                background-color: #2a82da;
                color: white;
                font-weight: bold;
            }
            QHeaderView::section {
                background-color: #555555;
                color: white;
                padding: 12px 8px;
                font-size: 13px;
                font-weight: bold;
                border: 1px solid #666666;
                text-align: center;
            }
            QHeaderView::section:hover {
                background-color: #666666;
            }
        """)
        
        # Hide row numbers and improve appearance
        self.live_table.verticalHeader().setVisible(False)
        self.live_table.setAlternatingRowColors(True)
        self.live_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        
        # Set optimal row height for better readability
        self.live_table.verticalHeader().setDefaultSectionSize(35)  # Taller rows for better readability
        
        # Set better column widths for optimal readability
        self.live_table.setColumnWidth(0, 50)   # Rank column - compact
        self.live_table.setColumnWidth(1, 250)  # Task column - wider for full task names
        self.live_table.setColumnWidth(2, 80)   # Value column 
        self.live_table.setColumnWidth(3, 80)   # Score column
        
        # Enable drag operations
        self.live_table.setDragEnabled(True)
        self.live_table.setDragDropMode(QAbstractItemView.DragDropMode.DragOnly)
        self.live_table.setDefaultDropAction(Qt.DropAction.CopyAction)
        self.live_table.setToolTip("💡 Drag tasks from this table to the calendar to schedule them!")
        
        results_layout.addWidget(self.live_table)
        
        # Export button (always visible)
        self.export_button = QPushButton('📊 Export to Excel')
        self.export_button.clicked.connect(self.export_to_excel)
        self.export_button.setToolTip("💾 Save your prioritized task list to Excel with calendar schedule")
        self.export_button.setStyleSheet("""
            QPushButton {
                background-color: #28a745;
                font-weight: bold;
                font-size: 12px;
                padding: 8px;
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #34ce57;
            }
        """)
        results_layout.addWidget(self.export_button)
        
        results_widget.setLayout(results_layout)
        
        # Add to splitter
        plot_splitter.addWidget(plot_widget)
        plot_splitter.addWidget(results_widget)
        plot_splitter.setSizes([250, 350])  # Give more space to results table (was 300, 200)
        
        left_layout.addWidget(plot_splitter)
        left_panel.setLayout(left_layout)
        
        # Right side: Calendar panel
        calendar_panel = self.create_calendar_panel()
        
        # Add both panels to main splitter
        main_splitter.addWidget(left_panel)
        main_splitter.addWidget(calendar_panel)
        main_splitter.setSizes([600, 400])  # Give more space to chart initially
        
        layout.addWidget(main_splitter)
        self.plot_panel.setLayout(layout)

    def create_calendar_panel(self):
        """Create the calendar panel for task scheduling"""
        calendar_widget = QWidget()
        calendar_layout = QVBoxLayout()
        
        # Calendar header
        calendar_header = QLabel("📅 Task Scheduler")
        calendar_header.setStyleSheet("color: #ffffff; font-weight: bold; font-size: 14px; padding: 10px;")
        calendar_layout.addWidget(calendar_header)
        
        # Instructions
        instructions = QLabel("💡 Drag tasks directly to any day")
        instructions.setStyleSheet("color: #888888; font-size: 11px; padding: 0 10px 5px 10px; font-style: italic;")
        calendar_layout.addWidget(instructions)
        
        # Calendar widget
        self.calendar = ScheduleCalendarWidget(self)
        self.calendar.setStyleSheet("""
            QCalendarWidget {
                background-color: #404040;
                color: white;
                border: 1px solid #555555;
                border-radius: 4px;
            }
            QCalendarWidget QAbstractItemView:enabled {
                background-color: #404040;
                color: white;
                selection-background-color: #2a82da;
            }
            QCalendarWidget QAbstractItemView:disabled {
                color: #888888;
            }
            QCalendarWidget QWidget {
                background-color: #404040;
                color: white;
            }
            QCalendarWidget QToolButton {
                color: white;
                background-color: #555555;
                border: 1px solid #666666;
                border-radius: 2px;
                margin: 2px;
            }
            QCalendarWidget QToolButton:hover {
                background-color: #2a82da;
            }
            QCalendarWidget QMenu {
                background-color: #404040;
                color: white;
                border: 1px solid #555555;
            }
            QCalendarWidget QSpinBox {
                background-color: #555555;
                color: white;
                border: 1px solid #666666;
                border-radius: 2px;
            }
        """)
        calendar_layout.addWidget(self.calendar)
        
        # Selected date display
        self.selected_date_label = QLabel("📍 Selected: Today")
        self.selected_date_label.setStyleSheet("""
            color: #2a82da; 
            font-weight: bold; 
            font-size: 12px; 
            padding: 8px; 
            background-color: #404040; 
            border-radius: 4px; 
            margin: 5px;
        """)
        calendar_layout.addWidget(self.selected_date_label)
        
        # Time selection
        time_frame = QGroupBox("⏰ Time Schedule")
        time_frame.setStyleSheet("""
            QGroupBox {
                color: white;
                font-weight: bold;
                border: 1px solid #555555;
                border-radius: 4px;
                margin-top: 10px;
                padding-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                subcontrol-position: top left;
                padding: 0 5px;
            }
        """)
        time_layout = QFormLayout()
        
        self.start_time = QTimeEdit()
        self.start_time.setTime(QTime(9, 0))  # Default 9:00 AM
        self.start_time.setStyleSheet("""
            QTimeEdit {
                background-color: #555555;
                color: white;
                border: 1px solid #666666;
                border-radius: 2px;
                padding: 4px;
            }
        """)
        
        self.end_time = QTimeEdit()
        self.end_time.setTime(QTime(10, 0))  # Default 10:00 AM
        self.end_time.setStyleSheet("""
            QTimeEdit {
                background-color: #555555;
                color: white;
                border: 1px solid #666666;
                border-radius: 2px;
                padding: 4px;
            }
        """)
        
        time_layout.addRow("Start:", self.start_time)
        time_layout.addRow("End:", self.end_time)
        time_frame.setLayout(time_layout)
        calendar_layout.addWidget(time_frame)
        
        # Scheduled tasks list
        scheduled_frame = QGroupBox("📋 Scheduled Tasks")
        scheduled_frame.setStyleSheet("""
            QGroupBox {
                color: white;
                font-weight: bold;
                border: 1px solid #555555;
                border-radius: 4px;
                margin-top: 10px;
                padding-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                subcontrol-position: top left;
                padding: 0 5px;
            }
        """)
        scheduled_layout = QVBoxLayout()
        
        self.scheduled_tasks_list = QListWidget()
        self.scheduled_tasks_list.setStyleSheet("""
            QListWidget {
                background-color: #404040;
                color: white;
                border: 1px solid #555555;
                border-radius: 4px;
                padding: 4px;
            }
            QListWidget::item {
                padding: 6px;
                border-bottom: 1px solid #555555;
            }
            QListWidget::item:selected {
                background-color: #2a82da;
            }
            QListWidget::item:hover {
                background-color: #505050;
            }
        """)
        self.scheduled_tasks_list.setMaximumHeight(150)
        scheduled_layout.addWidget(self.scheduled_tasks_list)
        
        # Clear schedule button
        clear_button = QPushButton("🗑️ Clear Selected")
        clear_button.clicked.connect(self.clear_selected_schedule)
        clear_button.setStyleSheet("""
            QPushButton {
                background-color: #dc3545;
                color: white;
                border: none;
                padding: 6px 12px;
                border-radius: 4px;
                font-size: 11px;
            }
            QPushButton:hover {
                background-color: #c82333;
            }
        """)
        scheduled_layout.addWidget(clear_button)
        
        scheduled_frame.setLayout(scheduled_layout)
        calendar_layout.addWidget(scheduled_frame)
        
        # Set up drag and drop
        self.calendar.setAcceptDrops(True)
        self.calendar.dragEnterEvent = self.calendar_drag_enter_event
        self.calendar.dragMoveEvent = self.calendar_drag_move_event
        self.calendar.dragLeaveEvent = self.calendar_drag_leave_event
        self.calendar.dropEvent = self.calendar_drop_event
        
        # Connect calendar date change to update scheduled tasks display
        self.calendar.selectionChanged.connect(self.update_scheduled_tasks_display)
        self.calendar.selectionChanged.connect(self.update_selected_date_display)
        
        # Connect month/year changes to refresh bold formatting
        self.calendar.currentPageChanged.connect(self.calendar.refresh_calendar_display)
        
        # Update the display initially
        self.update_selected_date_display()
        
        # Initial refresh to show any existing scheduled tasks in bold
        if hasattr(self, 'task_list') and self.task_list:
            self.calendar.refresh_calendar_display()
        
        calendar_widget.setLayout(calendar_layout)
        return calendar_widget

    def calendar_drag_enter_event(self, event):
        """Handle drag enter events for the calendar"""
        if event.mimeData().hasText() and event.mimeData().text().startswith("task_"):
            print("🎯 Drag started - entering calendar")
            self.calendar.drag_in_progress = True
            # Store original month to prevent navigation
            self.calendar.original_month = self.calendar.selectedDate()
            event.acceptProposedAction()
        else:
            event.ignore()

    def calendar_drag_move_event(self, event):
        """Handle drag move events for the calendar"""
        if event.mimeData().hasText() and event.mimeData().text().startswith("task_"):
            event.acceptProposedAction()
            
            # Get the date being hovered over and highlight it
            drop_pos = event.position()
            hovered_date = self.get_date_at_position(drop_pos)
            
            if hovered_date:
                # Only update highlighting if we're on a different date
                q_date = QDate(hovered_date.year, hovered_date.month, hovered_date.day)
                if not self.calendar.highlighted_date or self.calendar.highlighted_date != q_date:
                    self.calendar.highlight_date_for_drop(hovered_date)
                    
            # Prevent month navigation by restoring original month if it changed
            if hasattr(self.calendar, 'original_month'):
                current_month = self.calendar.selectedDate()
                if (current_month.year() != self.calendar.original_month.year() or 
                    current_month.month() != self.calendar.original_month.month()):
                    print("⚠️  Preventing month change during drag")
                    # Don't change the month, but allow day selection within the month
                    corrected_date = QDate(self.calendar.original_month.year(), 
                                         self.calendar.original_month.month(), 
                                         self.calendar.original_month.day())
                    self.calendar.setSelectedDate(corrected_date)
        else:
            # Clear highlighting if dragging something that's not a task
            self.calendar.clear_drop_highlighting()
            event.ignore()

    def calendar_drag_leave_event(self, event):
        """Handle drag leave events for the calendar"""
        print("🎯 Drag left calendar area")
        self.calendar.drag_in_progress = False
        self.calendar.clear_drop_highlighting()
        event.accept()  # Accept the event to ensure it's processed

    def calendar_drop_event(self, event):
        """Handle drop events for the calendar"""
        print("🎯 Drop event - cleaning up drag state")
        
        # Clear highlighting and drag state first
        self.calendar.drag_in_progress = False
        self.calendar.clear_drop_highlighting()
        
        if event.mimeData().hasText() and event.mimeData().text().startswith("task_"):
            # Extract task index from mime data
            task_index_str = event.mimeData().text().replace("task_", "")
            try:
                task_index = int(task_index_str)
                if 0 <= task_index < len(self.task_list):
                    # Get the date being dropped on (not just the selected date)
                    drop_pos = event.position()
                    
                    # Convert the drop position to a date
                    # We need to find which date cell the drop occurred on
                    calendar_date = self.get_date_at_position(drop_pos)
                    
                    if calendar_date:
                        task = self.task_list[task_index]
                        
                        # Show time selection dialog
                        dialog = TimeSelectionDialog(task.task, calendar_date, self)
                        
                        # Pre-fill with smart defaults
                        self.set_smart_default_times(dialog, calendar_date)
                        
                        if dialog.exec() == QDialog.DialogCode.Accepted:
                            start_time, end_time = dialog.get_times()
                            
                            # Schedule the task
                            task.schedule_on_calendar(
                                datetime.combine(calendar_date, datetime.min.time()),
                                start_time,
                                end_time
                            )
                            
                            # Update the calendar selection to the dropped date
                            q_date = QDate(calendar_date.year, calendar_date.month, calendar_date.day)
                            self.calendar.setSelectedDate(q_date)
                            
                            # Update the display
                            self.update_scheduled_tasks_display()
                            self.update_priority_display()
                            
                            # Refresh calendar to show bold formatting
                            self.calendar.refresh_calendar_display()
                            
                            event.acceptProposedAction()
                        else:
                            event.ignore()
                    else:
                        event.ignore()
                else:
                    event.ignore()
            except ValueError:
                event.ignore()
        else:
            event.ignore()

    def get_date_at_position(self, pos):
        """Get the date at the given position in the calendar widget - improved accuracy"""
        try:
            # Get calendar dimensions and current state
            calendar_rect = self.calendar.rect()
            current_date = self.calendar.selectedDate()
            
            # Debug the position
            print(f"🎯 Mouse position: ({pos.x():.1f}, {pos.y():.1f}) in calendar {calendar_rect.width()}x{calendar_rect.height()}")
            
            # Use a more reliable approach: try multiple header heights and see which works best
            possible_header_heights = [60, 80, 100, 120, 140]  # Try different header heights
            
            for header_height in possible_header_heights:
                # Calculate grid parameters
                content_height = calendar_rect.height() - header_height
                if content_height <= 0:
                    continue
                    
                cell_width = calendar_rect.width() / 7.0
                cell_height = content_height / 6.0
                
                # Calculate grid position with more precision
                grid_x = pos.x() / cell_width
                grid_y = (pos.y() - header_height) / cell_height
                
                # Check if we're in a reasonable grid position
                if 0 <= grid_x <= 7 and 0 <= grid_y <= 6:
                    col = int(grid_x)
                    row = int(grid_y)
                    
                    # Get calendar layout info
                    year = current_date.year()
                    month = current_date.month()
                    first_day = QDate(year, month, 1)
                    first_weekday = first_day.dayOfWeek() % 7  # Qt: Monday=1, convert to Sunday=0
                    
                    # Calculate the day number
                    day_number = row * 7 + col - first_weekday + 1
                    days_in_month = first_day.daysInMonth()
                    
                    print(f"  📅 Header: {header_height}, Grid: ({col}, {row}), Day: {day_number}")
                    
                    # If it's a valid day in the current month, use it
                    if 1 <= day_number <= days_in_month:
                        target_date = QDate(year, month, day_number)
                        if target_date.isValid():
                            result_date = datetime(target_date.year(), target_date.month(), target_date.day()).date()
                            print(f"  ✅ Found date: {result_date}")
                            return result_date
            
            # If none of the header heights worked, fall back to a simple approach
            print("  ⚠️  Falling back to selected date")
            
        except Exception as e:
            print(f"  ❌ Error in date calculation: {e}")
        
        # Fallback to current selected date
        current_selected = self.calendar.selectedDate()
        fallback_date = datetime(current_selected.year(), current_selected.month(), current_selected.day()).date()
        print(f"  🔄 Using fallback date: {fallback_date}")
        return fallback_date

    def set_smart_default_times(self, dialog, date):
        """Set smart default times based on existing scheduled tasks for the date"""
        # Find the latest end time for this date
        latest_end_time = QTime(9, 0)  # Default start at 9 AM
        
        for task in self.task_list:
            if (task.is_scheduled() and 
                task.scheduled_date.date() == date and 
                task.scheduled_end_time):
                try:
                    # Parse existing end time
                    task_end_time = QTime.fromString(task.scheduled_end_time, "hh:mm AP")
                    if not task_end_time.isValid():
                        task_end_time = QTime.fromString(task.scheduled_end_time, "HH:mm")
                    
                    if task_end_time.isValid() and task_end_time > latest_end_time:
                        latest_end_time = task_end_time
                except:
                    pass
        
        # If there are existing tasks, start after the latest one
        if latest_end_time != QTime(9, 0):
            # Add 30 minutes buffer
            start_time = latest_end_time.addSecs(30 * 60)
        else:
            start_time = QTime(9, 0)
        
        # Set end time 1 hour after start time
        end_time = start_time.addSecs(60 * 60)
        
        dialog.start_time.setTime(start_time)
        dialog.end_time.setTime(end_time)

    def update_scheduled_tasks_display(self):
        """Update the scheduled tasks list based on selected calendar date"""
        self.scheduled_tasks_list.clear()
        
        if not self.task_list:
            return
            
        q_date = self.calendar.selectedDate()
        selected_date = datetime(q_date.year(), q_date.month(), q_date.day()).date()
        
        for i, task in enumerate(self.task_list):
            if (task.is_scheduled() and 
                task.scheduled_date.date() == selected_date):
                
                time_info = ""
                if task.scheduled_start_time and task.scheduled_end_time:
                    time_info = f" ({task.scheduled_start_time} - {task.scheduled_end_time})"
                
                item_text = f"{task.task}{time_info}"
                list_item = QListWidgetItem(item_text)
                list_item.setData(Qt.ItemDataRole.UserRole, i)  # Store task index
                self.scheduled_tasks_list.addItem(list_item)

    def clear_selected_schedule(self):
        """Clear the schedule for the selected task"""
        current_item = self.scheduled_tasks_list.currentItem()
        if current_item:
            task_index = current_item.data(Qt.ItemDataRole.UserRole)
            if 0 <= task_index < len(self.task_list):
                self.task_list[task_index].clear_schedule()
                self.update_scheduled_tasks_display()
                self.update_priority_display()  # Refresh priority display
                
                # Refresh calendar to update bold formatting
                self.calendar.refresh_calendar_display()

    def add_test_goals_and_proceed(self):
        """Add test goals and automatically proceed to plot view"""
        test_goals = [
            ("Complete Project Proposal", 4.5, 3.0),
            ("Review Code Changes", 3.0, 2.0),
            ("Team Meeting", 2.5, 1.5),
            ("Update Documentation", 3.5, 4.0),
            ("Bug Fixing", 4.0, 2.5),
            ("Client Presentation", 5.0, 4.0),
            ("Code Refactoring", 3.5, 5.0),
            ("Unit Testing", 4.0, 3.0),
            ("Performance Optimization", 4.5, 6.0),
            ("Security Audit", 5.0, 4.5),
            ("Database Migration", 4.0, 7.0),
            ("API Integration", 3.5, 3.5),
            ("User Training", 3.0, 2.0),
            ("System Backup", 2.5, 1.0),
            ("Deployment Planning", 4.0, 2.0),
            ("Code Review", 3.5, 1.5),
            ("Feature Implementation", 4.5, 5.0),
            ("Technical Documentation", 3.0, 4.0),
            ("Bug Triage", 3.5, 2.0),
            ("System Monitoring", 2.5, 1.5)
        ]
        
        for task_name, value, time in test_goals:
            self.task_list.append(Task(task_name, value, time))
        
        self.refresh_input_table()
        self.proceed_to_plot()

    def add_goals_from_clipboard_and_proceed(self):
        """Add goals from clipboard and automatically proceed if successful"""
        clipboard = QApplication.clipboard()
        text = clipboard.text().strip()
        
        if not text:
            QMessageBox.warning(self, "📋 Clipboard Empty", "❌ No text found in clipboard!\n\n💡 Copy a list of tasks (one per line) and try again.")
            return
            
        goals = [goal.strip() for goal in text.split('\n') if goal.strip()]
        
        if not goals:
            QMessageBox.warning(self, "❌ No Valid Tasks", "The clipboard text doesn't contain valid tasks.\n\n💡 Make sure each task is on a separate line!")
            return
            
        for goal in goals:
            self.task_list.append(Task(goal, 3.0, 4.0))  # Default to middle of our ranges
            
        self.refresh_input_table()
        
        # Auto-proceed if we have a reasonable number of tasks
        if len(goals) >= 3:
            self.proceed_to_plot()
        else:
            QMessageBox.information(self, "✅ Tasks Added!", f"Added {len(goals)} tasks. Add more or click a task to start prioritizing!")

    def add_task_and_auto_proceed(self):
        """Add a task and auto-proceed to plot if we have enough tasks"""
        task = self.task_input.text().strip()
        if not task:
            QMessageBox.warning(self, "⚠️ Input Required", "🎯 Please enter a task name first!")
            return
        
        self.task_list.append(Task(task, 3.0, 4.0))  # Default to middle of our ranges
        self.task_input.clear()
        self.refresh_input_table()
        
        # Auto-proceed if we have 3 or more tasks
        if len(self.task_list) >= 3:
            self.proceed_to_plot()
        elif len(self.task_list) == 1:
            # Give encouraging feedback for first task
            self.task_input.setPlaceholderText("Great! Add 2 more tasks to start prioritizing...")

    def proceed_to_plot(self):
        """Smoothly transition to the plot view"""
        if not self.task_list:
            return
            
        # Show both panels
        self.plot_panel.show()
        self.main_splitter.setSizes([350, 650])  # Give more space to plot
        
        # Update plot and live results
        self.update_plot()
        self.update_priority_display()
        
        # Minimize input panel after a moment (optional)
        QTimer.singleShot(2000, lambda: self.main_splitter.setSizes([250, 750]))

    def update_priority_display(self):
        """Update the live priority ranking table"""
        if not self.task_list:
            return
            
        # Calculate scores
        for task in self.task_list:
            task.calculate_score()
            
        # Sort by score
        sorted_tasks = sorted(self.task_list, key=lambda t: t.score, reverse=True)
        
        # Update live table (show all tasks, but highlight top ones)
        display_count = min(15, len(sorted_tasks))  # Show more tasks (up to 15)
        self.live_table.setRowCount(display_count)
        
        for i, task in enumerate(sorted_tasks[:display_count]):
            # Rank with special indicators for top 3
            if i == 0:
                rank_text = "🥇"  # Gold medal for #1
            elif i == 1:
                rank_text = "🥈"  # Silver medal for #2  
            elif i == 2:
                rank_text = "🥉"  # Bronze medal for #3
            else:
                rank_text = f"#{i+1}"
                
            rank_item = QTableWidgetItem(rank_text)
            rank_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            if i < 3:  # Top 3 get special formatting
                font = rank_item.font()
                font.setBold(True)
                font.setPointSize(font.pointSize() + 1)
                rank_item.setFont(font)
            self.live_table.setItem(i, 0, rank_item)
            
            # Task name - now show FULL name since we have more space
            task_name = task.task  # Show full task name
            if task.is_scheduled():
                task_name = f"📅 {task_name}"  # Calendar icon for scheduled tasks
            
            task_item = QTableWidgetItem(task_name)
            task_item.setToolTip(f"Full task: {task.task}\nScheduled: {'Yes' if task.is_scheduled() else 'No'}")
            if i < 3:  # Top 3 get bold formatting
                font = task_item.font()
                font.setBold(True)
                task_item.setFont(font)
            self.live_table.setItem(i, 1, task_item)
            
            # Value with better formatting
            value_item = QTableWidgetItem(f"{task.value:.1f}⭐")
            value_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            value_item.setToolTip(f"Impact/Value rating: {task.value:.1f} out of 5.0")
            if i < 3:
                font = value_item.font()
                font.setBold(True)
                value_item.setFont(font)
            self.live_table.setItem(i, 2, value_item)
            
            # Score with color-coded priority indicators
            score_text = f"{task.score:.2f}"
            if task.score >= 2.0:
                score_text = f"🔥{score_text}"  # High priority
            elif task.score >= 1.0:
                score_text = f"⚡{score_text}"  # Medium priority
            else:
                score_text = f"💡{score_text}"  # Lower priority
                
            score_item = QTableWidgetItem(score_text)
            score_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            score_item.setToolTip(f"Priority Score: {task.score:.2f}\nCalculated as Value({task.value:.1f}) ÷ Time({task.time:.1f})")
            if i < 3:
                font = score_item.font()
                font.setBold(True)
                score_item.setFont(font)
            self.live_table.setItem(i, 3, score_item)
            
            # Enhanced highlighting for top 3 with gradient effect
            if i == 0:  # Gold highlighting for #1
                for col in range(4):
                    item = self.live_table.item(i, col)
                    if item:
                        item.setBackground(QColor(255, 215, 0, 150))  # Gold
                        item.setForeground(QColor(0, 0, 0))  # Black text on gold
            elif i == 1:  # Silver highlighting for #2
                for col in range(4):
                    item = self.live_table.item(i, col)
                    if item:
                        item.setBackground(QColor(192, 192, 192, 150))  # Silver
                        item.setForeground(QColor(0, 0, 0))  # Black text on silver
            elif i == 2:  # Bronze highlighting for #3
                for col in range(4):
                    item = self.live_table.item(i, col)
                    if item:
                        item.setBackground(QColor(205, 127, 50, 150))  # Bronze
                        item.setForeground(QColor(255, 255, 255))  # White text on bronze
        
        # Don't auto-resize columns since we set fixed widths for optimal readability
        # self.live_table.resizeColumnsToContents()  # Commented out to keep our fixed widths

    def refresh_input_table(self):
        self.input_table.setRowCount(len(self.task_list))
        for i, t in enumerate(self.task_list):
            self.input_table.setItem(i, 0, QTableWidgetItem(t.task))

    def on_press(self, event):
        if event.inaxes != self.ax:
            return
        contains, ind = self.scatter.contains(event)
        if contains:
            task_index = ind["ind"][0]
            
            # Left-click: normal dragging within chart
            if event.button == 1:  # Left mouse button
                self.dragging = True
                self.drag_index = task_index

    def on_motion(self, event):
        if not self.dragging or self.drag_index is None or event.inaxes != self.ax:
            return
        
        # Update task values
        self.task_list[self.drag_index].value = max(0, min(6, event.xdata))  # Clamp to valid range
        self.task_list[self.drag_index].time = max(0, min(8, event.ydata))   # Clamp to valid range
        self.moved_points.add(self.drag_index)  # Mark this point as moved
        
        # Update scatter plot data directly for smooth movement
        x_data = [t.value for t in self.task_list]
        y_data = [t.time for t in self.task_list]
        self.scatter.set_offsets(np.column_stack([x_data, y_data]))
        
        # Update the annotation position
        for i, annotation in enumerate(self.ax.texts):
            if i == self.drag_index:
                annotation.set_position((event.xdata, event.ydata))
        
        self.canvas.draw_idle()  # Use draw_idle for smoother updates
        
        # Trigger real-time priority update with a small delay
        self.auto_update_timer.start(100)  # Update after 100ms of no movement

    def on_release(self, event):
        if self.dragging:
            # Do a full redraw when drag is complete
            self.update_plot()
            # Update priority display immediately
            self.update_priority_display()
        self.dragging = False
        self.drag_index = None

    def on_hover(self, event):
        if event.inaxes != self.ax:
            if self.current_annotation:
                self.current_annotation.set_visible(False)
                self.current_annotation = None
                self.canvas.draw_idle()
            return

        contains, ind = self.scatter.contains(event)
        if contains:
            pos = ind["ind"][0]
            task = self.task_list[pos]
            
            # Remove previous annotation if it exists
            if self.current_annotation:
                self.current_annotation.set_visible(False)
            
            # Create new annotation with task name and values
            priority_score = task.value / task.time if task.time > 0 else 0
            text = f"{task.task}\nValue: {task.value:.1f}\nTime: {task.time:.1f}\nPriority: {priority_score:.2f}"
            self.current_annotation = self.ax.annotate(
                text,
                xy=(task.value, task.time),
                xytext=(10, 10),
                textcoords='offset points',
                bbox=dict(
                    boxstyle='round,pad=0.5',
                    fc='#2a82da',
                    ec='#555555',
                    alpha=0.9
                ),
                color='white',
                fontsize=9,
                fontweight='bold',
                arrowprops=dict(
                    arrowstyle='->',
                    connectionstyle='arc3,rad=0.2',
                    color='#555555',
                    linewidth=1.5
                )
            )
            self.canvas.draw_idle()
        elif self.current_annotation:
            self.current_annotation.set_visible(False)
            self.current_annotation = None
            self.canvas.draw_idle()

    def update_plot(self):
        self.ax.clear()
        
        # Reapply modern styling
        self.ax.set_facecolor('#353535')
        self.ax.grid(True, linestyle='--', alpha=0.3, color='#555555')
        self.ax.spines['bottom'].set_color('#555555')
        self.ax.spines['top'].set_color('#555555')
        self.ax.spines['left'].set_color('#555555')
        self.ax.spines['right'].set_color('#555555')
        
        # Set labels with modern styling
        self.ax.set_xlabel('* Value (Impact/Importance)', color='white', fontsize=11, fontweight='bold')
        self.ax.set_ylabel('Time Investment (Hours)', color='white', fontsize=11, fontweight='bold')
        self.ax.set_title('Interactive Priority Matrix', color='white', fontsize=13, fontweight='bold', pad=15)
        
        # Style the ticks
        self.ax.tick_params(colors='white', which='both')
        
        # Maintain fixed axis limits
        self.ax.set_xlim(0, 6)
        self.ax.set_ylim(0, 8)
        
        # Create arrays for all points
        x_data = [t.value for t in self.task_list]
        y_data = [t.time for t in self.task_list]
        
        # Calculate scores and find the top 3 tasks
        for task in self.task_list:
            task.calculate_score()
        
        # Get indices sorted by score (highest first)
        sorted_indices = sorted(range(len(self.task_list)), key=lambda i: self.task_list[i].score, reverse=True)
        top_3_indices = sorted_indices[:3]  # Get top 3 tasks
        
        # Create color array based on whether points have been moved and scheduled
        colors = []
        for i in range(len(self.task_list)):
            if self.task_list[i].is_scheduled():
                colors.append('#28a745')  # Green for scheduled tasks
            elif i in self.moved_points:
                colors.append('#2a82da')  # Blue for moved but unscheduled
            else:
                colors.append('#e74c3c')  # Red for original position unscheduled
        
        # Create scatter plot for all points except the top 3
        non_top_indices = [i for i in range(len(self.task_list)) if i not in top_3_indices]
        if non_top_indices:
            self.ax.scatter(
                [x_data[i] for i in non_top_indices],
                [y_data[i] for i in non_top_indices],
                c=[colors[i] for i in non_top_indices],
                picker=True,
                alpha=0.7,
                s=100
            )
        
        # Plot the top 3 tasks with circled numbers
        for rank, task_index in enumerate(top_3_indices, 1):
            if task_index < len(self.task_list):  # Safety check
                task_x = x_data[task_index]
                task_y = y_data[task_index]
                self.ax.plot(task_x, task_y, 'o', markersize=20, markerfacecolor='none', 
                            markeredgecolor=colors[task_index], markeredgewidth=2)
                self.ax.text(task_x, task_y, str(rank), ha='center', va='center', 
                            fontsize=14, fontweight='bold', color=colors[task_index])
        
        # Update the scatter reference for event handling
        self.scatter = self.ax.scatter(x_data, y_data, c=colors, picker=True, alpha=0)
        
        # Adjust figure layout
        self.figure.subplots_adjust(left=0.15, bottom=0.15, right=0.95, top=0.9)
        self.canvas.draw()

    def export_to_excel(self):
        if not self.task_list:
            QMessageBox.warning(self, "❌ No Data", "There are no tasks to export!\n\n🎯 Add some tasks and prioritize them first.")
            return
            
        # Get save file path
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "💾 Save Priority Report",
            f"priority_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Excel Files (*.xlsx)"
        )
        
        if not file_path:
            return
            
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Task Priorities"
            
            # Add headers including calendar information
            headers = ['📋 Task', '★ Value', '⏰ Time (hours)', '🏆 Priority Score', 
                      '📅 Scheduled Date', '🕐 Start Time', '🕐 End Time']
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Add data
            sorted_tasks = calculate_and_sort_tasks(self.task_list)
            for row, task in enumerate(sorted_tasks, 2):
                ws.cell(row=row, column=1, value=task.task)
                ws.cell(row=row, column=2, value=task.value)
                ws.cell(row=row, column=3, value=task.time)
                ws.cell(row=row, column=4, value=task.score)
                
                # Add calendar scheduling information
                if task.is_scheduled():
                    ws.cell(row=row, column=5, value=task.scheduled_date.strftime('%Y-%m-%d'))
                    ws.cell(row=row, column=6, value=task.scheduled_start_time if task.scheduled_start_time else "")
                    ws.cell(row=row, column=7, value=task.scheduled_end_time if task.scheduled_end_time else "")
                else:
                    ws.cell(row=row, column=5, value="Not scheduled")
                    ws.cell(row=row, column=6, value="")
                    ws.cell(row=row, column=7, value="")
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column[0].column_letter].width = adjusted_width
            
            wb.save(file_path)
            QMessageBox.information(self, "✅ Export Successful!", f"🎉 Your priority analysis has been saved with calendar schedule!\n\n📁 File location: {file_path}\n\n💡 You can now share this with your team or use it for planning.")
            
        except Exception as e:
            QMessageBox.critical(self, "❌ Export Error", f"😞 Failed to export your data:\n\n{str(e)}\n\n💡 Try saving to a different location or check file permissions.")

    def show_help(self):
        help_text = """
🎯 <b>Welcome to PriPlot - Your Task Priority Assistant!</b>

<b>🚀 Quick Start Options:</b>
• "🧪 Try Sample Tasks" - Instantly explore with 20 realistic work tasks
• "📋 Import List" - Paste tasks from your clipboard (one per line)
• Type manually and press Enter to add tasks one by one

<b>📊 Interactive Prioritization:</b>
• Once you have 3+ tasks, the plot automatically appears
• Left-click and drag tasks on chart to set their Value (→) and Time Investment (↑)
• Bottom-right corner = High Value + Low Time = TOP PRIORITY!
• See live priority rankings update as you drag

<b>📅 Calendar Scheduling:</b>
• Drag tasks from the priority ranking table directly to any day on the calendar
• A time selection dialog will appear showing the day you dropped on
• Choose your preferred start and end times in the dialog
• Scheduled tasks appear in green on the chart and with 📅 in rankings
• Days with scheduled tasks appear in bold on the calendar
• View scheduled tasks by clicking different calendar dates
• Remove schedules using the "Clear Selected" button

<b>🏆 Understanding Your Results:</b>
• Priority Score = Value ÷ Time (higher is better)
• Top 3 tasks get special numbered circles on the plot
• Live ranking table shows your current priorities with schedule indicators
• Export to Excel includes both priorities and calendar schedules

<b>💡 Pro Tips:</b>
• Hover over plot points to see task details and priority scores
• Color coding: Red = original, Blue = repositioned, Green = scheduled
• Focus on high-value, low-time tasks for maximum impact
• Schedule your top priority tasks first for better time management

<b>🎨 What the Numbers Mean:</b>
• Value (0-6): How important/impactful is this task?
• Time (0-8): How many hours will this take?
• Priority Score: Automatically calculated as Value ÷ Time

Happy prioritizing and scheduling! 🚀📅
        """
        
        msg = QMessageBox(self)
        msg.setWindowTitle("📚 PriPlot User Guide")
        msg.setText(help_text)
        msg.setStyleSheet("""
            QMessageBox {
                background-color: #353535;
                color: white;
            }
            QMessageBox QLabel {
                color: white;
                font-size: 13px;
                min-width: 500px;
            }
        """)
        msg.exec()

    def show_welcome_message(self):
        welcome_text = """
🎉 <b>Welcome to PriPlot!</b>

Transform your task management with smart priority visualization!

<b>🚀 New Streamlined Experience:</b>
• Choose "🧪 Try Sample Tasks" for instant exploration
• Or "📋 Import List" to paste your own tasks
• Add 3+ tasks and the plot appears automatically
• See live priority updates as you drag tasks around

<b>💡 No more clicking through tabs!</b>
Everything flows naturally as you work. The interface adapts to show you exactly what you need, when you need it.

Ready to boost your productivity? 🎯
        """
        
        msg = QMessageBox(self)
        msg.setWindowTitle("🎯 Welcome to PriPlot!")
        msg.setText(welcome_text)
        msg.setStyleSheet("""
            QMessageBox {
                background-color: #353535;
                color: white;
            }
            QMessageBox QLabel {
                color: white;
                font-size: 13px;
                min-width: 400px;
            }
            QMessageBox QPushButton {
                min-width: 80px;
                padding: 8px;
            }
        """)
        msg.exec()

    def update_selected_date_display(self):
        """Update the selected date display label"""
        if hasattr(self, 'selected_date_label'):
            selected_date = self.calendar.selectedDate().toString("dddd, MMMM d, yyyy")
            self.selected_date_label.setText(f"📍 Selected: {selected_date}") 